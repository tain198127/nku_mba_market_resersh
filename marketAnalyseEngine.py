import logging
import math
import os
import random
import sys

import matplotlib.pyplot as plt
import numpy
import openpyxl
import xlrd
from sklearn import cluster

logging.basicConfig(level=logging.DEBUG)
family_name = ['赵', '钱', '孙', '李', '周', '吴', '郑', '王', '冯', '陈', '褚', '卫', '蒋', '沈', '韩', '杨', '朱', '秦', '尤', '许',
               '何', '吕', '施', '张', '孔', '曹', '严', '华', '金', '魏', '陶', '姜', '戚', '谢', '邹', '喻', '柏', '水', '窦', '章',
               '云', '苏', '潘', '葛', '奚', '范', '彭', '郎', '鲁', '韦', '昌', '马', '苗', '凤', '花', '方', '俞', '任', '袁', '柳',
               '酆', '鲍', '史', '唐', '费', '廉', '岑', '薛', '雷', '贺', '倪', '汤', '滕', '殷', '罗', '毕', '郝', '邬', '安', '常',
               '乐', '于', '时', '傅', '皮', '卞', '齐', '康', '伍', '余', '元', '卜', '顾', '孟', '平', '黄', '和', '穆', '萧', '尹',
               '姚', '邵', '堪', '汪', '祁', '毛', '禹', '狄', '米', '贝', '明', '臧', '计', '伏', '成', '戴', '谈', '宋', '茅', '庞',
               '熊', '纪', '舒', '屈', '项', '祝', '董', '梁']

first_names = [
    '子', '中', '生', '国', '年', '和',
    '会', '家', '可', '天', '去', '能', '多', '然', '于', '心', '学', '都',
    '看', '发', '当', '成', '只', '如', '事', '还', '用', '第', '样', '道', '想', '作', '种', '开', '美',
    '总', '从', '己', '面', '前', '些', '同', '日', '又', '行', '意', '动',
    '方', '期', '头', '经', '长', '儿', '回', '位', '分', '爱', '因', '名', '法', '间', '斯', '知',
    '世', '者', '高', '已', '亲', '其', '进', '此', '常', '与', '活', '正', '感',
    '见', '明', '问', '力', '理', '尔', '点', '文', '几', '定', '本', '公', '特', '外', '相', '西', '果', '走',
    '将', '月', '十', '实', '向', '声', '车', '全', '信', '重', '三', '机', '工', '物', '气', '真',
    '太', '新', '比', '才', '夫', '再', '书', '部', '水', '体', '却', '加', '电', '门',
    '海', '听', '表', '德', '少', '克', '代', '员', '许', '稜', '先', '安',
    '光', '白', '住', '望', '教', '命', '花', '结', '乐', '东', '神', '记', '处', '让',
    '直', '平', '报', '友', '关', '放', '至', '入', '笑', '内', '英', '军',
    '候', '民', '岁', '往', '何', '度', '山', '觉', '路', '带', '万', '男', '边', '风', '解', '任', '金', '快', '原',
    '通', '师', '立', '象', '数', '四', '失', '满', '战', '远', '格', '士', '音', '轻', '目',
    '始', '达', '深', '完', '今', '提', '求', '清', '王', '化', '空', '业', '思', '切', '非', '找', '罗',
    '钱', '语', '元', '喜', '曾', '离', '飞', '科', '言', '干', '流', '欢', '约', '各', '即', '指', '合',
    '题', '必', '该', '论', '交', '终', '林', '医', '晚', '决', '窢', '传', '画', '保', '运',
    '则', '房', '早', '院', '量', '火', '布', '品', '近', '坐', '产', '答', '星', '精', '视', '五', '连', '司', '巴',
    '奇', '管', '类', '未', '朋', '台', '夜', '青', '北', '久', '乎', '越', '观', '落', '尽', '形', '影',
    '红', '百', '令', '周', '识', '步', '希', '亚', '术', '留', '市', '半', '热', '送', '兴', '造', '谈', '容',
    '极', '随', '演', '收', '首', '根', '讲', '整', '式', '取', '照', '办', '强', '石', '古', '华', '諣', '拿', '计',
    '足', '双', '米', '丽', '客', '南', '领', '节', '衣', '站', '黑', '刻', '统',
    '福', '城', '故', '历', '选', '包', '争', '另', '建', '维', '绝', '树', '系', '示', '愿',
    '持', '千', '史', '准', '联', '纪', '基', '买', '志', '静', '阿', '诗', '独', '复', '消', '社', '算',
    '义', '竟', '确', '单', '治', '卡', '幸', '兰', '念', '举', '钟', '共', '毛', '句', '息', '功',
    '官', '待', '究', '跟', '穿', '室', '易', '游', '程', '号', '居', '考', '突', '皮', '费', '价', '图', '具',
    '刚', '脑', '永', '歌', '响', '商', '礼', '细', '专', '黄', '块', '味', '灵', '改', '据', '般', '破', '引', '食',
    '存', '众', '注', '笔', '甚', '某', '沉', '备', '习', '校', '默', '务', '土', '微', '须', '试', '怀',
    '调', '广', '蜖', '苏', '显', '赛', '查', '密', '议', '底', '列', '富', '梦', '参', '八',
    '亮', '印', '线', '温', '京', '初', '养', '香', '停', '际', '致', '阳', '纸', '纳', '验',
    '助', '激', '够', '严', '证', '忘', '趣', '支', '春', '集', '丈', '木', '研', '班', '普', '导', '顿',
    '展', '获', '艺', '六', '波', '察', '群', '段', '庭', '创', '区', '奥', '器', '谢', '弟', '店',
    '草', '排', '背', '止', '组', '州', '朝', '封', '睛', '板', '角', '况', '曲', '馆', '育', '质', '河', '续',
    '呼', '若', '推', '境', '遇', '雨', '标', '充', '围', '案', '伦', '护', '冷', '警', '贝', '著', '雪', '索',
    '剧', '船', '险', '烟', '依', '斗', '值', '帮', '汉', '佛', '肯', '闻', '唱', '沙', '局', '伯', '族', '低',
    '玩', '资', '屋', '击', '速', '顾', '泪', '洲', '团', '圣', '旁', '堂', '兵', '七', '露', '园', '牛', '旅',
    '劳', '型', '烈', '姑', '陈', '莫', '鱼', '异', '抱', '宝', '权', '鲁', '简', '态', '级', '票', '怪', '寻', '律',
    '胜', '汽', '右', '洋', '范', '舞', '秘', '午', '登', '楼', '贵', '责', '例', '追', '较', '职', '属',
    '渐', '左', '录', '丝', '牙', '党', '继', '托', '赶', '章', '智', '冲', '叶', '胡', '吉', '坚', '遗',
    '修', '松', '临', '藏', '担', '戏', '善', '卫', '药', '悲', '敢', '伊', '村', '戴', '词', '森', '耳',
    '祖', '云', '规', '散', '迷', '油', '适', '乡', '恩', '投', '弹', '铁', '博', '雷', '府',
    '超', '勒', '杂', '醒', '洗', '采', '毫', '毕', '九', '冰', '既', '状', '乱', '景', '席', '珍', '童', '顶',
    '派', '素', '脱', '农', '疑', '练', '野', '按', '征', '骨', '余', '承', '置', '臓', '彩', '灯', '巨',
    '琴', '环', '技', '束', '增', '忍', '洛', '忆', '判', '欧', '层', '付',
    '阵', '玛', '批', '岛', '项', '休', '懂', '武', '革', '良', '恶', '恋', '委', '拥', '娜', '妙', '探', '呀', '营',
    '退', '摇', '弄', '桌', '熟', '诺', '宣', '银', '势', '奖', '宫', '忽', '套', '康', '供', '优', '课', '鸟',
    '夏', '健', '模', '伴', '守', '挥', '鲜', '财', '孤', '枪', '杰',
    '迹', '遍', '盖', '坦', '江', '顺', '秋', '萨', '授', '归', '浪', '凡', '预',
    '雄', '升', '典', '莱', '含', '盛', '济', '蒙', '棋', '端', '释', '介', '烧',
    '乾', '坤']


class MarketAnalyseEngine:
    @staticmethod
    def random_name():
        """
        获取一个随机名字
        :return: 随机姓名
        """
        name = random.choice(family_name)
        name = name + random.choice(first_names)
        name_idx = random.randint(0, 1)
        if name_idx > 0:
            name = name + random.choice(first_names)
        return name

    @staticmethod
    def __open_table(excel_path, sheet_idx):
        """
        :param excel_path: xlsx的地址
        :param sheet_idx:  第几个sheet，从0开始
        :return: sheet对象
        """
        data = xlrd.open_workbook(excel_path)
        sheet = data.sheets()[sheet_idx]
        return sheet

    def __read_personal_info(self, excel_path):
        """
        读取个人信息数组
        :param excel_path: 文件路径
        :return:一份个人信息数组
        """
        table = self.__open_table(excel_path, 0)
        gander = table.cell_value(0, 1)
        age = table.cell_value(1, 1)
        location = table.cell_value(2, 1)
        edu = table.cell_value(3, 1)
        marital = table.cell_value(4, 1)
        industry = table.cell_value(5, 1)
        career = table.cell_value(6, 1)
        lowest_consumption = table.cell_value(7, 1)
        highest_comsumption = table.cell_value(7, 3)
        device = table.cell_value(8, 1)
        info = [gander, age, location, edu, marital, industry, career, lowest_consumption, highest_comsumption, device]
        logging.debug("begin--------------" + sys._getframe().f_code.co_name)
        logging.debug(info)
        logging.debug("end--------------" + sys._getframe().f_code.co_name)
        return info

    def __read_detail_info(self, excel_path):
        """
        读取详细信息矩阵
        :param excel_path: 文件路径
        :return: 一份详细信息矩阵
        """
        detail = []
        table = self.__open_table(excel_path, 1)
        rows_count = 0
        column_count = 48
        real_column_count = 0
        """
        表明表头是两行
        """
        first_col = table.col_values(0)
        first_row = table.row_values(2)
        real_column_count = len(first_row)
        for row in first_col:
            if row != '-':
                rows_count = rows_count + 1
        logging.debug(rows_count)
        # 48列
        for row in range(2, rows_count):
            row_values = table.row_values(row)
            detail.append(row_values)

        logging.debug(detail)
        # 删掉第一行
        detail = numpy.delete(detail, 1, axis=1)
        logging.debug(detail)
        # 删掉后面几列
        detail = numpy.delete(detail, range(47, real_column_count - 1), axis=1)
        logging.debug("begin--------------" + sys._getframe().f_code.co_name)
        logging.debug(detail)
        logging.debug("end--------------" + sys._getframe().f_code.co_name)
        return detail

    def __read_assemble_info(self, excel_path):
        """
        读取汇总信息
        :param excel_path:  excel 文件路径
        :return: 一份汇总信息
        """
        table = self.__open_table(excel_path, 2)
        info = table.row_values(2)
        logging.debug("begin--------------" + sys._getframe().f_code.co_name)
        logging.debug(info)
        logging.debug("end--------------" + sys._getframe().f_code.co_name)
        return info

    @staticmethod
    def read_excels():
        """
        读取文件
        :return:统计文档的文件名列表
        """
        excels = []
        logging.debug("begin--------------" + sys._getframe().f_code.co_name)
        logging.debug(os.getcwd())
        logging.debug("end--------------" + sys._getframe().f_code.co_name)
        path = os.getcwd()
        for root, dirs, files in os.walk(path):
            for file in files:
                logging.debug("filename:[{}]".format(file))
                logging.debug("ext:[{}]".format(os.path.splitext(file)[1]))
                if os.path.splitext(file)[1] == ".xlsx" and os.path.splitext(file)[0] not in [
                    'user_behaviour_statistics', 'asm']:
                    excel_path = os.path.join(root, file)
                    excels.append(excel_path)
        return excels

    @staticmethod
    def __writ_into_excel(file_name, person_info_ary, detail_matrix_ary, asm_info_ary):
        """
        写入excel
        :param file_name: 文件路径
        :param person_info_ary: 个人信息数组
        :param detail_matrix_ary: 个人使用app的详情矩阵的数组
        :param asm_info_ary: 个人的汇总信息数组
        :return: 保存的文件路径
        """
        excel_file = os.path.join(os.path.dirname(os.getcwd()), file_name)
        if os.path.exists(excel_file):
            os.remove(excel_file)
        logging.debug(excel_file)
        logging.debug(person_info_ary)
        logging.debug(detail_matrix_ary)
        logging.debug(asm_info_ary)

        wb = openpyxl.Workbook()
        sheet1 = wb.create_sheet(title="personal info", index=0)
        sheet1.append(["name", "gander", "age", "local", "edu", "marital", "industry", "career", "lowest_consumption",
                       "highest_comsumption", "devie"])
        for person_info in person_info_ary:
            sheet1.append(person_info)
        sheet2 = wb.create_sheet(title="detail info", index=1)
        sheet2.append(
            ['姓名', 'app名称', '6:00~9:00', '9:00~12:00', '12:00~14:00', '14:00~19:00', '19:00~23:00', '23:00~6:00', '在家',
             '上班路上-公交',
             '上班路上-私家车', '旅游', '办公室', '出差-短途', '出差-飞机', '出差-高铁等', '晴天', '阴天', '刮风', '下雨', '下雪', '雾霾', '台风', '工作日',
             '	周末', '	春节', '	国庆节', '	劳动节', '	清明	', '端午', '	情人节', '	中秋节', '	元旦', '	圣诞节',
             '	万圣节', '	体育', '	娱乐', '	旅游', '	房产', '	汽车	', '美食', '	理财	', '邮件网络', '	社群网络',
             '	平均每周使用多少次', '	平均为此APP付款金额(元/周)	', '平均每周使用时长(分钟/周)'])
        for detail_matrix in detail_matrix_ary:
            for detail_info in detail_matrix:
                sheet2.append(detail_info.tolist())
        sheet3 = wb.create_sheet(title="assembly info", index=2)
        sheet3.append(['姓名', '6:00~9:00', '9:00~12:00', '12:00~14:00', '14:00~19:00', '19:00~23:00', '23:00~6:00', '在家',
                       '上班路上-公交',
                       '上班路上-私家车', '旅游', '办公室', '出差-短途', '出差-飞机', '出差-高铁等', '晴天', '阴天', '刮风', '下雨', '下雪', '雾霾', '台风',
                       '工作日',
                       '	周末', '	春节', '	国庆节', '	劳动节', '	清明	', '端午', '	情人节', '	中秋节', '	元旦',
                       '	圣诞节',
                       '	万圣节', '	体育', '	娱乐', '	旅游', '	房产', '	汽车	', '美食', '	理财	'])
        for asm_info in asm_info_ary:
            sheet3.append(asm_info)
        wb.save(excel_file)
        return excel_file

    def _getMergeMatrix(self, is_normalize=False):
        """
        获取汇总矩阵
        :param is_normalize: 是否要做一般化处理
        :return: 个人信息矩阵、详情矩阵、汇总矩阵
        """
        excels = self.read_excels()
        personal_merge_info = []
        detail_merge_info = []
        assemb_merge_info = []
        for f in excels:
            random_name = self.random_name()
            personal_info = self.__read_personal_info(f)
            detail_info = self.__read_detail_info(f)
            assemb_info = self.__read_assemble_info(f)
            # 写入随机姓名
            personal_info.insert(0, random_name)
            detail_info = numpy.insert(detail_info, 0, values=random_name, axis=1)
            assemb_info.insert(0, random_name)
            # 打印日志
            logging.debug(personal_info)
            logging.debug(assemb_info)
            logging.debug(detail_info)
            personal_merge_info.append(personal_info)
            detail_merge_info.append(detail_info)
            assemb_merge_info.append(assemb_info)
        if is_normalize:
            return self._normalization(personal_merge_info, detail_merge_info, assemb_merge_info)
        else:
            return personal_merge_info, detail_merge_info, assemb_merge_info

    def merge_into_excel(self, file_name, is_normalization):
        """
        主函数，合并脚本所在目录的素有xlsx文件，集成了读取、汇总和写入操作
        :return: 合并后的文件地址
        """

        personal_merge_info, detail_merge_info, assemb_merge_info = self._getMergeMatrix(is_normalization)
        merge_file_path = self.__writ_into_excel(file_name, personal_merge_info, detail_merge_info, assemb_merge_info)
        return merge_file_path

    def _normalization(self, person, detail, asm):
        """
        一般化处理
        :param person: 个人信息矩阵
        :param detail:  详情矩阵
        :param asm: 统计矩阵
        :return: 详情矩阵做0-1化，统计矩阵做百分比化
        """
        nor_detail = numpy.array(detail)
        nor_detail[nor_detail == '是'] = 1
        nor_detail[nor_detail == '否'] = 0
        nor_detail[nor_detail == ''] = 0
        # 清洗detail
        logging.debug(nor_detail)

        nor_asm = []
        for row in asm:
            name = numpy.split(row, [1, 7])[0]
            schedule = numpy.split(row, [1, 7])[1].astype(float)
            surrounding = numpy.split(row, [7, 15])[1].astype(float)
            weather = numpy.split(row, [15, 22])[1].astype(float)
            holiday = numpy.split(row, [22, 34])[1].astype(float)
            classification = numpy.split(row, [34, 41])[1].astype(float)
            # 先转换为fload
            nor_row = numpy.hstack((name,
                                    schedule / numpy.sum(schedule),
                                    surrounding / numpy.sum(surrounding),
                                    weather / numpy.sum(weather),
                                    holiday / numpy.sum(holiday),
                                    classification / numpy.sum(classification))
                                   )
            # numpy实现，百分比化
            nor_asm.append(nor_row.tolist())

        return person, nor_detail, nor_asm

    def kMeanCluster(self):
        """
        K-MEAN分类
        :return:
        """
        person, detail, asm = self._getMergeMatrix(True)
        # logging.debug(asm)
        data = numpy.asmatrix(asm)[:, 1:7]
        logging.debug(data)

        # print(data)
        for k in range(2, math.ceil(data.shape[1] / 2)):
            numSample = len(data)
            centroid, label, inertia = cluster.k_means(data, k)
            logging.debug(centroid, label, inertia)
            mark = ['or', 'ob', 'og', 'ok', '^r', '+r', 'sr', 'dr', '<r', 'pr']
            for i in range(numSample):
                plt.plot(data[i][0], data[i][0], mark[label[i]])
            mark = ['Dr', 'Db', 'Dg', 'Dk', '^b', '+b', 'sb', 'db', '<b', 'pb']
            for i in range(k):
                plt.plot(centroid[i][0], centroid[i][1], mark[label[i]], markersize=12)
            plt.show()
        # a = numpy.array([10, 11, 9, 23, 21, 11, 45, 20, 11, 12]).reshape(-1, 1)
        # kde = KernelDensity(kernel='gaussian', bandwidth=3).fit(a)
        # s = numpy.linspace(0, 50)
        # e = kde.score_samples(s.reshape(-1, 1))
        # plt.plot(s, e)
        # plt.show()


engin = MarketAnalyseEngine()
for i in range(1,100):
    print(engin.random_name())
# engin.kMeanCluster()
# docs = engin.merge_into_excel("asm.xlsx",True)
